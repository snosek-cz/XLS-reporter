#!/usr/bin/env python3
"""
generate_test_data.py
=====================
Generates 12 months of dummy time-tracking XLSX files for the XLS Reporter project.

Each file contains a single 'Details MONTH YEAR' sheet mimicking data from
a consulting company (Deloitte UK style), populated with Blackadder-inspired
team members and realistic development task descriptions.

Output: test_data/Details_MONTH_YEAR.xlsx  (12 files, Apr 2026 – Mar 2027)

Usage:
    python generate_test_data.py
"""

import openpyxl
import datetime
import os
import random
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

random.seed(42)  # reproducible output

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), 'test_data')
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─── TEAM DEFINITION ──────────────────────────────────────────────────────────
# (name, role, rate, pod, wbs, base_hours_per_day, available_month_indices)
TEAM = [
    ('Edmund Blackadder',   'Senior Manager',       185, 'Cobra POD',  'GBL0007741.2.11', 6, list(range(12))),
    ('S. Baldrick',         'Junior Developer',      95, 'Cobra POD',  'GBL0007741.2.11', 8, list(range(12))),
    ('George St. Barleigh', 'Senior Developer',     130, 'Cobra POD',  'GBL0007741.2.11', 8, list(range(12))),
    ('Nurse Darling',       'UX/UI Lead',           140, 'Viper POD',  'GBL0007741.2.22', 7, list(range(12))),
    ('Gen. Melchett',       'Cloud Architect',      210, 'Viper POD',  'GBL0007741.2.22', 5, [0,1,2,3,5,6,7,8,9,10,11]),  # absent month 4
    ('Capt. Kevin Darling', 'QA Test Lead',         120, 'Viper POD',  'GBL0007741.2.22', 7, list(range(12))),
    ('Bobbie Parkhurst',    'Full Stack Developer', 125, 'Python POD', 'GBL0007741.2.33', 8, [0,1,2,3,4,5,6,7,9,10,11]),  # absent month 8
    ('Doris Miggins',       'Data Engineer',        135, 'Python POD', 'GBL0007741.2.33', 8, [1,2,3,4,5,6,7,8,9,10,11]),  # joins month 1
    ('Lord Percy Percy',    'DevOps Engineer',      145, 'Python POD', 'GBL0007741.2.33', 7, list(range(12))),
    ('Mrs. Miggins',        'Business Analyst',     115, 'Cobra POD',  'GBL0007741.2.11', 6, [0,1,2,3,4,5,8,9,10,11]),    # absent months 6,7
]

CLIENT = 'Deloitte UK'

PROJECT_DESCRIPTIONS = {
    'Cobra POD': [
        'Requirements gathering - stakeholder interviews',
        'Sprint planning and backlog grooming',
        'User story refinement with client team',
        'Cunning plan implementation - phase 2',
        'Stakeholder presentation preparation',
        'Risk register update and mitigation planning',
        'Client status report and dashboard update',
        'Architecture review board preparation',
        'Cunning plan - revised after client feedback',
        'Change request assessment and documentation',
        'Vendor evaluation and procurement support',
        'Project governance documentation',
    ],
    'Viper POD': [
        'UI component library development',
        'Figma wireframes review and iteration',
        'Cloud infrastructure design - AWS',
        'CI/CD pipeline configuration',
        'Security review and penetration testing prep',
        'UAT test case preparation',
        'Regression testing suite - automated',
        'UX research synthesis',
        'A/B testing framework setup',
        'Accessibility audit - WCAG 2.1 compliance',
        'Design system documentation',
        'Performance testing and optimisation',
    ],
    'Python POD': [
        'API development - REST endpoints',
        'Data pipeline implementation',
        'Database schema optimisation',
        'Microservices refactoring',
        'Infrastructure as code - Terraform',
        'Kubernetes cluster configuration',
        'Data quality framework implementation',
        'ETL pipeline debugging and tuning',
        'Code review and PR management',
        'Technical debt remediation',
        'Monitoring and alerting setup - Datadog',
        'Documentation - Confluence update',
    ],
}

BALDRICK_DESCRIPTIONS = [
    'Cunning plan implementation - stage 1',
    'Cunning plan implementation - stage 2 (revised)',
    'Cunning plan - the one involving a turnip',
    'Cunning plan debugging - turnip not working',
    'Cunning plan v3 - considerably less cunning',
    'Support ticket resolution - something went wrong',
    'Code review feedback implementation',
    'Unit tests - mostly passing',
    'Meeting - did not understand most of it',
    'Documentation - added some words',
]

MELCHETT_DESCRIPTIONS = [
    'Cloud architecture - MAGNIFICENT design review',
    'Technical strategy - bold and daring approach',
    'Architecture decision record - brilliant insight',
    'Cloud cost optimisation - slashing costs ruthlessly',
    'Executive briefing - delivered with great authority',
    'AWS Well-Architected review - passed with flying colours',
    'Security posture review - ZERO tolerance for weakness',
]

BLACKADDER_DESCRIPTIONS = [
    'Stakeholder management - kept them from bothering real team',
    'Executive reporting - made bad news sound acceptable',
    'Risk management - identified risk of client noticing delays',
    'Status update - technically accurate if selectively read',
    'Scope negotiation - cunning redefinition of "done"',
    'Budget review - numbers adjusted for palatability',
    'Team motivation - threatened mild inconveniences',
    'Governance compliance - ticked all the boxes',
]

FEATURES = ['Authentication', 'Dashboard', 'Reporting', 'API Integration',
            'Data Migration', 'Infrastructure', 'Security', 'Performance',
            'User Management', 'Analytics', 'Notifications', '']
PBIS = ['PBI-' + str(i) for i in range(1001, 1050)] + ['']


def get_working_days(year, month):
    """Return list of working days (Mon-Fri) for the given month."""
    import calendar
    days = []
    for d in range(1, calendar.monthrange(year, month)[1] + 1):
        dt = datetime.date(year, month, d)
        if dt.weekday() < 5:
            days.append(dt)
    return days


def iso_week(dt):
    return dt.isocalendar()[1]


# ─── MONTHS: April 2026 – March 2027 ──────────────────────────────────────────
MONTHS = [(2026, m) for m in range(4, 13)] + [(2027, m) for m in range(1, 4)]

print(f'Generating {len(MONTHS)} monthly files into: {OUTPUT_DIR}')
print()

for month_idx, (year, month) in enumerate(MONTHS):
    month_name = datetime.date(year, month, 1).strftime('%B')
    sheet_name = f'Details {month_name} {year}'
    fname = f'Details_{month_name}_{year}.xlsx'
    fpath = os.path.join(OUTPUT_DIR, fname)

    working_days = get_working_days(year, month)

    # One person takes leave during the 2nd ISO week each month (rotating)
    vacation_person = TEAM[month_idx % len(TEAM)][0]
    vacation_week = sorted(set(iso_week(d) for d in working_days))[1]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # ── Header row ──────────────────────────────────────────────────────────
    headers = ['Client', 'Name', 'Date', 'Project', 'WBS', 'Project role',
               'Feature', 'PBI', 'Description', 'Logged time (hours)',
               'Rate', 'Rate per MD/per hour', 'Total fee', 'Week num']

    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10, name='Calibri')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30

    # ── Data rows ────────────────────────────────────────────────────────────
    row = 2
    for person_name, role, rate, pod, wbs, base_hours, avail_months in TEAM:
        if month_idx not in avail_months:
            continue

        if 'Baldrick' in person_name:
            descs = BALDRICK_DESCRIPTIONS
        elif 'Melchett' in person_name:
            descs = MELCHETT_DESCRIPTIONS
        elif 'Blackadder' in person_name:
            descs = BLACKADDER_DESCRIPTIONS
        else:
            descs = PROJECT_DESCRIPTIONS[pod]

        for day in working_days:
            wk = iso_week(day)
            if person_name == vacation_person and wk == vacation_week:
                continue
            if 'Senior Manager' in role and random.random() < 0.15:
                continue
            if 'Architect' in role and random.random() < 0.10:
                continue

            hours = max(1, min(10, base_hours + random.randint(-2, 1)))
            desc = random.choice(descs)
            feature = random.choice(FEATURES)
            pbi = random.choice(PBIS) if feature else ''

            ws.cell(row=row, column=1, value=CLIENT)
            ws.cell(row=row, column=2, value=person_name)
            date_cell = ws.cell(row=row, column=3, value=datetime.datetime(day.year, day.month, day.day))
            date_cell.number_format = 'dd\.mm\.yyyy'
            ws.cell(row=row, column=4, value=pod)
            ws.cell(row=row, column=5, value=wbs)
            ws.cell(row=row, column=6, value=role)
            ws.cell(row=row, column=7, value=feature)
            ws.cell(row=row, column=8, value=pbi)
            ws.cell(row=row, column=9, value=desc)
            ws.cell(row=row, column=10, value=hours)
            ws.cell(row=row, column=11, value=rate)
            ws.cell(row=row, column=12, value='per hour')
            ws.cell(row=row, column=13, value=hours * rate)
            ws.cell(row=row, column=14, value=f'=_xlfn.ISOWEEKNUM(C{row})')
            row += 1

    # ── Column widths ────────────────────────────────────────────────────────
    col_widths = [15, 22, 14, 18, 22, 22, 18, 12, 45, 10, 8, 18, 12, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Zebra stripes ────────────────────────────────────────────────────────
    fill_light = PatternFill(start_color='EBF3FB', end_color='EBF3FB', fill_type='solid')
    data_font = Font(size=10, name='Calibri')
    for r in range(2, row):
        for c in range(1, 15):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            if r % 2 == 0:
                cell.fill = fill_light

    wb.save(fpath)
    print(f'  ✅ {fname:40s} ({row - 2} rows)')

print()
print('Done!')
