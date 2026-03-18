#!/usr/bin/env python3
"""
consolidate_annual.py
=====================
Consolidates processed monthly XLSX files into a single Annual_YEAR.xlsx workbook.

Scans a folder for Details_MONTH_YEAR.xlsx files that contain both a Details
sheet and an Overview sheet, then copies all sheets into Annual_YEAR.xlsx.
Months are ordered chronologically. Re-running updates existing months.

Usage:
    python consolidate_annual.py /path/to/Done/folder
    python consolidate_annual.py /path/to/Done/folder --notify   # macOS notification
    python consolidate_annual.py /path/to/Done/folder --year 2026 # specific year only
"""

import sys
import os
import re
import datetime
import glob
import shutil
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

# Month name → number mapping
MONTH_NUM = {
    'January':1,'February':2,'March':3,'April':4,'May':5,'June':6,
    'July':7,'August':8,'September':9,'October':10,'November':11,'December':12
}


def parse_filename(fname):
    """
    Parse 'Details_April_2026.xlsx' → ('April', 2026)
    Returns (month_name, year) or (None, None) if not matching.
    """
    m = re.match(r'^Details_([A-Za-z]+)_(\d{4})\.xlsx$', os.path.basename(fname))
    if m:
        return m.group(1), int(m.group(2))
    return None, None


def copy_sheet(source_wb, source_sheet_name, target_wb, target_sheet_name):
    """
    Copy a worksheet from source_wb into target_wb with a new name.
    Copies values, formulas, basic styles and column widths.
    """
    src = source_wb[source_sheet_name]

    # Remove existing sheet with same name in target
    if target_sheet_name in target_wb.sheetnames:
        del target_wb[target_sheet_name]

    # Create new sheet
    dst = target_wb.create_sheet(title=target_sheet_name)

    # Copy column widths
    for col_letter, col_dim in src.column_dimensions.items():
        dst.column_dimensions[col_letter].width = col_dim.width

    # Copy row heights
    for row_num, row_dim in src.row_dimensions.items():
        dst.row_dimensions[row_num].height = row_dim.height

    # Copy cells
    for row in src.iter_rows():
        for cell in row:
            new_cell = dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.alignment = copy(cell.alignment)
                new_cell.border = copy(cell.border)
                new_cell.number_format = cell.number_format

    return dst


def consolidate(folder, year_filter=None, notify=False):
    """
    Scan folder for processed monthly files and build Annual_YEAR.xlsx for each year found.
    """
    folder = os.path.abspath(folder)
    print(f'Scanning: {folder}')

    # Find all processed monthly files
    candidates = glob.glob(os.path.join(folder, 'Details_*.xlsx'))

    # Group by year
    by_year = {}
    for fpath in candidates:
        month_name, year = parse_filename(fpath)
        if month_name is None:
            continue
        if year_filter and year != year_filter:
            continue
        if month_name not in MONTH_NUM:
            continue
        # Only process files that have an overview sheet (i.e. already processed)
        try:
            wb_check = load_workbook(fpath, read_only=True, data_only=False)
            overview_name = f'{month_name} {year}'
            details_name  = f'Details {month_name} {year}'
            has_overview = overview_name in wb_check.sheetnames
            has_details  = details_name in wb_check.sheetnames
            wb_check.close()
            if not has_overview or not has_details:
                print(f'  ⚠️  Skipping {os.path.basename(fpath)} — missing overview sheet (run generate_overview.py first)')
                continue
        except Exception as e:
            print(f'  ⚠️  Could not open {os.path.basename(fpath)}: {e}')
            continue

        by_year.setdefault(year, []).append((MONTH_NUM[month_name], month_name, year, fpath))

    if not by_year:
        print('  No processable files found.')
        return []

    annual_files = []

    for year in sorted(by_year.keys()):
        months = sorted(by_year[year], key=lambda x: x[0])  # sort by month number
        annual_path = os.path.join(folder, f'Annual_{year}.xlsx')

        print(f'\n  Building Annual_{year}.xlsx ({len(months)} months)...')

        # Load or create annual workbook
        if os.path.exists(annual_path):
            annual_wb = load_workbook(annual_path)
            print(f'    Updating existing Annual_{year}.xlsx')
        else:
            annual_wb = openpyxl.Workbook()
            # Remove default sheet
            if 'Sheet' in annual_wb.sheetnames:
                del annual_wb['Sheet']
            print(f'    Creating new Annual_{year}.xlsx')

        # Track which sheets we want (in order)
        desired_sheets = []  # [(overview_name, details_name), ...]
        for month_num, month_name, year, fpath in months:
            desired_sheets.append((
                f'{month_name} {year}',
                f'Details {month_name} {year}',
                fpath
            ))

        # Copy sheets from each monthly file into annual workbook
        for overview_name, details_name, fpath in desired_sheets:
            src_wb = load_workbook(fpath)
            print(f'    + {overview_name}')
            copy_sheet(src_wb, overview_name, annual_wb, overview_name)
            copy_sheet(src_wb, details_name, annual_wb, details_name)
            src_wb.close()

        # Reorder sheets chronologically
        desired_order = []
        for overview_name, details_name, _ in desired_sheets:
            desired_order.append(overview_name)
            desired_order.append(details_name)

        # Move any extra sheets (not in desired order) to end
        extra_sheets = [s for s in annual_wb.sheetnames if s not in desired_order]
        final_order = desired_order + extra_sheets

        # Apply sheet order
        for i, sheet_name in enumerate(final_order):
            if sheet_name in annual_wb.sheetnames:
                annual_wb.move_sheet(sheet_name, offset=i - annual_wb.sheetnames.index(sheet_name))

        annual_wb.save(annual_path)
        print(f'    ✅ Saved: {annual_path}')
        annual_files.append(annual_path)

        if notify:
            _macos_notify(
                title='XLS Reporter',
                message=f'Annual_{year}.xlsx updated ({len(months)} months)',
                subtitle=annual_path
            )

    return annual_files


def _macos_notify(title, message, subtitle=''):
    import subprocess
    script = f'display notification "{message}" with title "{title}" subtitle "{subtitle}"'
    subprocess.run(['osascript', '-e', script], capture_output=True)


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Usage: python consolidate_annual.py <folder> [--notify] [--year YYYY]')
        sys.exit(1)

    folder  = sys.argv[1]
    notify  = '--notify' in sys.argv
    year_filter = None
    if '--year' in sys.argv:
        idx = sys.argv.index('--year')
        if idx + 1 < len(sys.argv):
            year_filter = int(sys.argv[idx + 1])

    try:
        annual_files = consolidate(folder, year_filter=year_filter, notify=notify)
        if annual_files:
            print(f'\nDone. {len(annual_files)} annual file(s) created/updated.')
        else:
            print('\nNothing to consolidate.')
    except Exception as e:
        print(f'❌ Error: {e}', file=sys.stderr)
        if notify:
            _macos_notify('XLS Reporter — Error', str(e))
        sys.exit(1)
