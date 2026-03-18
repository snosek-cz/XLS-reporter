#!/usr/bin/env python3
"""
generate_overview.py
====================
Generates (or regenerates) an overview/summary sheet from a Details sheet
in an XLS Reporter XLSX workbook.

Workflow:
  1. Receives a Details_MONTH_YEAR.xlsx file path
  2. Reads the 'Details MONTH YEAR' sheet
  3. Aggregates hours by (Name, POD, WBS, Rate) per week-of-month
  4. Writes a 'MONTH YEAR' overview sheet matching the template style
  5. Saves the file in place

Usage:
    python generate_overview.py /path/to/Details_April_2026.xlsx
    python generate_overview.py /path/to/Details_April_2026.xlsx --notify  # macOS notification
"""

import sys
import os
import datetime
from collections import defaultdict
from copy import copy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.styles.numbers import FORMAT_NUMBER
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle


# ─── CONSTANTS ────────────────────────────────────────────────────────────────

# Column indices in Details sheet (1-based)
DET_NAME      = 2   # B - Resource name
DET_DATE      = 3   # C - Date
DET_PROJECT   = 4   # D - POD / Project name
DET_WBS       = 5   # E - WBS code
DET_HOURS     = 10  # J - Logged time (hours)
DET_RATE      = 11  # K - Rate

# Column indices in Overview sheet (1-based)
OV_START_COL  = 2   # B - first data column
OV_NAME       = 2   # B
OV_POD        = 3   # C
OV_WBS        = 4   # D
OV_RATE       = 5   # E
OV_WEEK_START = 6   # F - Week 1 starts here
OV_HEADER_ROW = 2
OV_DATA_START = 3


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def iso_week(dt):
    return dt.isocalendar()[1]


def find_details_sheet(wb):
    """Find the first sheet whose name starts with 'Details'."""
    for name in wb.sheetnames:
        if name.startswith('Details'):
            return wb[name], name
    return None, None


def overview_sheet_name(details_name):
    """'Details April 2026' → 'April 2026'"""
    return details_name[len('Details '):].strip()


def get_weeks_ordered(ws_detail):
    """
    Return ISO weeks sorted by first occurrence date in the sheet.
    These map to Week 1, Week 2, ... Week N.
    """
    week_first_date = {}
    for row in ws_detail.iter_rows(min_row=2, values_only=True):
        date_val = row[DET_DATE - 1]
        if not date_val:
            continue
        if not isinstance(date_val, (datetime.datetime, datetime.date)):
            continue
        dt = date_val if isinstance(date_val, datetime.datetime) else datetime.datetime.combine(date_val, datetime.time())
        w = iso_week(dt)
        if w not in week_first_date:
            week_first_date[w] = dt
    return sorted(week_first_date.keys(), key=lambda w: week_first_date[w])


def aggregate_hours(ws_detail, weeks_ordered):
    """
    Aggregate hours by (name, pod, wbs, rate) per week.
    Returns ordered list of (key, {iso_week: hours}) preserving first-seen order.
    """
    hours = defaultdict(lambda: defaultdict(float))
    key_order = []
    seen_keys = set()

    for row in ws_detail.iter_rows(min_row=2, values_only=True):
        name = row[DET_NAME - 1]
        date_val = row[DET_DATE - 1]
        pod = row[DET_PROJECT - 1]
        wbs = row[DET_WBS - 1]
        logged = row[DET_HOURS - 1]
        rate = row[DET_RATE - 1]

        if not name or not date_val or not logged:
            continue
        if not isinstance(date_val, (datetime.datetime, datetime.date)):
            continue

        dt = date_val if isinstance(date_val, datetime.datetime) else datetime.datetime.combine(date_val, datetime.time())
        w = iso_week(dt)
        key = (name, pod, wbs, rate)

        if key not in seen_keys:
            key_order.append(key)
            seen_keys.add(key)

        hours[key][w] += float(logged)

    return key_order, hours


# ─── STYLING ──────────────────────────────────────────────────────────────────

def make_header_font():
    return Font(bold=True, size=12, name='Aptos Narrow')

def make_data_font(bold=False):
    return Font(bold=bold, size=12, name='Aptos Narrow')

def make_header_fill():
    # Matches template: theme color 2 (light blue-grey heading)
    from openpyxl.styles.fills import PatternFill
    from openpyxl.styles.colors import Color
    fill = PatternFill(fill_type='solid')
    fill.fgColor = Color(theme=2, type='theme')
    return fill

def make_total_fill():
    from openpyxl.styles.fills import PatternFill
    from openpyxl.styles.colors import Color
    fill = PatternFill(fill_type='solid')
    fill.fgColor = Color(theme=4, tint=0.6, type='theme')
    return fill


# ─── MAIN GENERATION ──────────────────────────────────────────────────────────

def generate_overview(filepath, notify=False):
    print(f'Processing: {filepath}')

    if not os.path.exists(filepath):
        raise FileNotFoundError(f'File not found: {filepath}')

    wb = openpyxl.load_workbook(filepath)

    # Find Details sheet
    ws_detail, details_name = find_details_sheet(wb)
    if ws_detail is None:
        raise ValueError('No Details sheet found in workbook.')

    print(f'  Details sheet: "{details_name}"')

    ov_name = overview_sheet_name(details_name)
    print(f'  Overview sheet: "{ov_name}"')

    # Remove existing overview sheet if present
    if ov_name in wb.sheetnames:
        del wb[ov_name]
        print(f'  Removed existing overview sheet.')

    # Get week ordering and aggregated data
    weeks_ordered = get_weeks_ordered(ws_detail)
    num_weeks = len(weeks_ordered)
    print(f'  Weeks in month: {num_weeks} (ISO: {weeks_ordered})')

    key_order, hours = aggregate_hours(ws_detail, weeks_ordered)
    print(f'  People/rows: {len(key_order)}')

    # Create overview sheet (insert before Details sheet)
    detail_idx = wb.sheetnames.index(details_name)
    ws_ov = wb.create_sheet(title=ov_name, index=detail_idx)

    # ── Row 1: empty ──────────────────────────────────────────────────────────
    # (left blank as per template)

    # ── Row 2: Headers ────────────────────────────────────────────────────────
    header_font = make_header_font()
    header_fill = make_header_fill()

    headers = ['Resource Name', 'POD Name', 'Work Package Code', 'Rate']
    headers += [f'Week {i}' for i in range(1, num_weeks + 1)]
    headers += ['Total Hours', 'Fee', 'Group PM']

    for i, h in enumerate(headers):
        col = OV_START_COL + i
        cell = ws_ov.cell(row=OV_HEADER_ROW, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    # ── Data rows ─────────────────────────────────────────────────────────────
    total_col    = OV_WEEK_START + num_weeks      # K column equivalent
    fee_col      = total_col + 1                  # L
    grouppm_col  = fee_col + 1                    # M

    row = OV_DATA_START
    for key in key_order:
        name, pod, wbs, rate = key

        ws_ov.cell(row=row, column=OV_NAME,  value=name).font  = make_data_font(bold=True)
        ws_ov.cell(row=row, column=OV_POD,   value=pod).font   = make_data_font()
        ws_ov.cell(row=row, column=OV_WBS,   value=wbs).font   = make_data_font()
        ws_ov.cell(row=row, column=OV_RATE,  value=rate).font  = make_data_font()

        # Week columns
        for w_idx, iso_w in enumerate(weeks_ordered):
            col = OV_WEEK_START + w_idx
            h = hours[key].get(iso_w, None)
            cell = ws_ov.cell(row=row, column=col, value=h)
            cell.font = make_data_font()

        # Total Hours formula: =SUM(F{row}:{week_end_col}{row})
        week_end_col_letter = get_column_letter(OV_WEEK_START + num_weeks - 1)
        week_start_col_letter = get_column_letter(OV_WEEK_START)
        total_cell = ws_ov.cell(row=row, column=total_col)
        total_cell.value = f'=SUM({week_start_col_letter}{row}:{week_end_col_letter}{row})'
        total_cell.font = make_data_font()

        # Fee formula: =TotalCol{row}*RateCol{row}
        total_col_letter = get_column_letter(total_col)
        rate_col_letter  = get_column_letter(OV_RATE)
        fee_cell = ws_ov.cell(row=row, column=fee_col)
        fee_cell.value = f'={total_col_letter}{row}*{rate_col_letter}{row}'
        fee_cell.font = make_data_font()

        row += 1

    # ── TOTAL row ─────────────────────────────────────────────────────────────
    total_row = row
    first_data_row = OV_DATA_START
    last_data_row  = row - 1

    total_font = make_data_font(bold=True)
    total_fill = make_total_fill()

    # Label
    label_cell = ws_ov.cell(row=total_row, column=OV_NAME, value='TOTAL')
    label_cell.font = total_font
    label_cell.fill = total_fill

    # Week totals
    for w_idx in range(num_weeks):
        col = OV_WEEK_START + w_idx
        col_letter = get_column_letter(col)
        cell = ws_ov.cell(row=total_row, column=col)
        cell.value = f'=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})'
        cell.font = total_font
        cell.fill = total_fill

    # Total hours total
    tc_letter = get_column_letter(total_col)
    total_hours_cell = ws_ov.cell(row=total_row, column=total_col)
    total_hours_cell.value = f'=SUM({tc_letter}{first_data_row}:{tc_letter}{last_data_row})'
    total_hours_cell.font = total_font
    total_hours_cell.fill = total_fill

    # Fee total
    fc_letter = get_column_letter(fee_col)
    fee_total_cell = ws_ov.cell(row=total_row, column=fee_col)
    fee_total_cell.value = f'=SUM({fc_letter}{first_data_row}:{fc_letter}{last_data_row})'
    fee_total_cell.font = total_font
    fee_total_cell.fill = total_fill

    # Apply total fill to empty cells in row for visual consistency
    for col in [OV_POD, OV_WBS, OV_RATE, grouppm_col]:
        cell = ws_ov.cell(row=total_row, column=col)
        cell.fill = total_fill

    # ── Column widths (match template style) ──────────────────────────────────
    col_widths = {
        1: 3,    # A - empty margin
        2: 24,   # B - Resource Name
        3: 18,   # C - POD Name
        4: 22,   # D - WBS
        5: 8,    # E - Rate
    }
    for w_idx in range(num_weeks):
        col_widths[OV_WEEK_START + w_idx] = 10  # Week columns
    col_widths[total_col]   = 13  # Total Hours
    col_widths[fee_col]     = 12  # Fee
    col_widths[grouppm_col] = 14  # Group PM

    for col, width in col_widths.items():
        ws_ov.column_dimensions[get_column_letter(col)].width = width

    # ── Save ──────────────────────────────────────────────────────────────────
    wb.save(filepath)
    print(f'  ✅ Saved: {filepath}')

    if notify:
        _macos_notify(
            title='XLS Reporter',
            message=f'Overview generated: {ov_name}',
            subtitle=os.path.basename(filepath)
        )

    return ov_name, len(key_order), num_weeks


def _macos_notify(title, message, subtitle=''):
    """Show a macOS notification via osascript."""
    import subprocess
    script = f'display notification "{message}" with title "{title}" subtitle "{subtitle}"'
    subprocess.run(['osascript', '-e', script], capture_output=True)


# ─── ENTRY POINT ──────────────────────────────────────────────────────────────

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Usage: python generate_overview.py <path_to_xlsx> [--notify]')
        sys.exit(1)

    filepath = sys.argv[1]
    notify   = '--notify' in sys.argv

    try:
        ov_name, n_people, n_weeks = generate_overview(filepath, notify=notify)
        print(f'  Overview "{ov_name}" created ({n_people} rows, {n_weeks} weeks)')
    except Exception as e:
        print(f'  ❌ Error: {e}', file=sys.stderr)
        if notify:
            _macos_notify('XLS Reporter — Error', str(e))
        sys.exit(1)
